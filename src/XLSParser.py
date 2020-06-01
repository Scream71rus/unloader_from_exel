
import openpyxl


class XLSParser:
    def __init__(self, parse_map, normalizer):
        self.parse_map = parse_map
        self.parsed_sheet_map = {}
        self.normalizer = normalizer

    def parser(self, filename):
        sheets = openpyxl.load_workbook(filename=filename)
        candidates = []

        for sheet in sheets:
            new_map = {}
            for row in sheet.iter_rows(max_row=1):
                for k, v in self.parse_map.items():
                    for column in row:
                        if column.value == v.get('title'):
                            new_map[k] = {**v, 'title': column.column}

            self.parsed_sheet_map[sheet.title] = new_map

            for row in sheet.iter_rows(min_row=2):
                data = {}
                for k, v in self.parsed_sheet_map.get(sheet.title).items():
                    for column in row:
                        if column.column == v.get('title'):
                            data[k] = self.normalizer.normalizers(v.get('normalize'))(str(column.value).strip())

                candidates.append(data)

        return candidates

    def run(self, filename):
        return self.parser(filename)
